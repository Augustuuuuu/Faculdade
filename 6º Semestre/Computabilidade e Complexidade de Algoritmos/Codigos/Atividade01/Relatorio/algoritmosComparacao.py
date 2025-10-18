import time
import platform
import psutil
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# pip install openpyxl psutil py-cpuinfo reportlab

# Tenta importar as dependências necessárias
try:
    import cpuinfo
except ImportError:
    cpuinfo = None

try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
except ImportError:
    print("ERRO: A biblioteca 'reportlab' não foi encontrada. Por favor, instale-a com: pip install reportlab")
    exit()

# --- Funções de Benchmark ---

def somar_linear(n):
    """Calcula a soma de 1 até n. Complexidade O(n)."""
    soma = 0
    for i in range(1, n + 1):
        soma += i
    return soma

def somar_quadratico(n):
    """Calcula a soma dos quadrados de 1 até n. Complexidade O(n)."""
    soma = 0
    for i in range(1, n + 1):
        soma += i**2
    return soma

def reducao_logaritmica(n):
    """Calcula quantos passos são necessários para reduzir n a 1 por divisão inteira. Complexidade O(log n)."""
    if n < 1:
        return 0
    contador_passos = 0
    i = n
    while i > 1:
        i = i // 2
        contador_passos += 1
    return contador_passos

# --- Funções de Sistema e Utilitários ---

def get_info_sistema():
    """Captura as especificações do sistema e retorna como um dicionário."""
    try:
        info = {
            "Sistema Operacional": f"{platform.system()} {platform.release()} ({platform.version()})",
            "Processador": cpuinfo.get_cpu_info().get('brand_raw', 'Não identificado') if cpuinfo else platform.processor(),
            "Núcleos Físicos": str(psutil.cpu_count(logical=False)),
            "Núcleos Lógicos": str(psutil.cpu_count(logical=True)),
            "Memória Total (GB)": f"{round(psutil.virtual_memory().total / (1024**3), 2)}"
        }
        return info
    except Exception as e:
        return {"Erro ao obter informações": str(e)}

def salvar_resultados_excel(*blocos_de_teste):
    """Salva os resultados em uma nova aba de uma planilha Excel."""
    filename = "resultados_benchmark.xlsx"
    # (O código desta função permanece o mesmo da versão anterior)
    try:
        workbook = openpyxl.load_workbook(filename) if os.path.exists(filename) else openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames and workbook.active.max_row == 1:
            if workbook.active.cell(1,1).value is None: workbook.remove(workbook.active)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        sheet = workbook.create_sheet(f"Resultados_{timestamp}")
        headers = ["Teste N°", "Passos (n)", "TEMPO DE EXECUÇÃO (s)"]
        for titulo, dados in blocos_de_teste:
            if sheet.max_row > 1: sheet.append([])
            sheet.merge_cells(start_row=sheet.max_row + 1, start_column=1, end_row=sheet.max_row + 1, end_column=len(headers))
            cell_titulo = sheet.cell(row=sheet.max_row, column=1, value=titulo)
            cell_titulo.font = Font(bold=True, size=14)
            sheet.append(headers)
            for cell in sheet[sheet.max_row]: cell.font = Font(bold=True)
            for resultado in dados: sheet.append(resultado)
        for col in sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            sheet.column_dimensions[column_letter].width = max_length + 2
        workbook.save(filename)
        print(f"\nResultados salvos com sucesso na aba '{sheet.title}' do arquivo '{filename}'")
    except Exception as e:
        print(f"Ocorreu um erro ao salvar a planilha Excel: {e}")


def salvar_resultados_pdf(info_sistema, *blocos_de_teste):
    """Gera um relatório em PDF com as informações do sistema e os resultados dos benchmarks."""
    filename = "relatorio_benchmark.pdf"
    doc = SimpleDocTemplate(filename)
    styles = getSampleStyleSheet()
    story = []

    # Título do Documento
    story.append(Paragraph("Relatório de Benchmark", styles['h1']))
    story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 0.25*inch))

    # Tabela com Informações do Sistema
    story.append(Paragraph("Especificações do Sistema", styles['h2']))
    info_data = [ [Paragraph(k, styles['Normal']), Paragraph(v, styles['Normal'])] for k, v in info_sistema.items() ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.25*inch))

    # Tabelas com Resultados dos Testes
    for titulo, dados in blocos_de_teste:
        story.append(Paragraph(titulo, styles['h2']))
        
        # Prepara os dados para a tabela, convertendo tudo para string
        tabela_dados = [["TESTE N°", "INPUT (n)", "TEMPO (s)"]]
        for linha in dados:
            # Formata o tempo para 6 casas decimais
            tabela_dados.append([str(linha[0]), str(linha[1]), f"{linha[2]:.6f}"])

        results_table = Table(tabela_dados)
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(results_table)
        story.append(Spacer(1, 0.25*inch))

    try:
        doc.build(story)
        print(f"Relatório em PDF salvo com sucesso no arquivo '{filename}'")
    except Exception as e:
        print(f"Ocorreu um erro ao gerar o PDF: {e}")


# --- Programa Principal ---
try:
    qtde_testes_input = input("Quantos testes deseja realizar para cada função? (Padrão: 5) ")
    qtde_testes = int(qtde_testes_input) if qtde_testes_input else 5
except ValueError:
    print("Entrada inválida. Usando o valor padrão de 5 testes.")
    qtde_testes = 5

# Coleta e exibe as informações do sistema
info_pc = get_info_sistema()
print("\n--- ESPECIFICAÇÕES DO SISTEMA ---")
for key, value in info_pc.items():
    print(f"{key}: {value}")
print("---------------------------------\n")

# --- 1. TESTES DE SOMA LINEAR ---
print("--- INICIANDO TESTES: SOMA LINEAR (O(n)) ---")
dados_linear = []
for teste in range(1, qtde_testes + 1):
    n = teste * 10000
    inicio = time.perf_counter()
    somar_linear(n)
    fim = time.perf_counter()
    tempo = fim - inicio
    print(f"Teste {teste}: n={n} | Tempo: {tempo:.6f}s")
    dados_linear.append([teste, n, tempo])

# --- 2. TESTES DE SOMA QUADRÁTICA ---
print("\n--- INICIANDO TESTES: SOMA QUADRÁTICA (O(n)) ---")
dados_quadratico = []
for teste in range(1, qtde_testes + 1):
    n = teste * 10000
    inicio = time.perf_counter()
    somar_quadratico(n)
    fim = time.perf_counter()
    tempo = fim - inicio
    print(f"Teste {teste}: n={n} | Tempo: {tempo:.6f}s")
    dados_quadratico.append([teste, n, tempo])

# --- 3. TESTES DE REDUÇÃO LOGARÍTMICA ---
print("\n--- INICIANDO TESTES: REDUÇÃO LOGARÍTMICA (O(log n)) ---")
dados_logaritmico = []
for teste in range(1, qtde_testes + 1):
    n = (teste * 100000) ** 10 
    inicio = time.perf_counter()
    reducao_logaritmica(n)
    fim = time.perf_counter()
    tempo = fim - inicio
    print(f"Teste {teste}: n≈{n:.0e} | Tempo: {tempo:.6f}s")
    # Corrigido: 'tempo' é adicionado como float, sem o "s"
    dados_logaritmico.append([teste, f"{n:.0e}", tempo])

# --- 4. SALVAR TUDO EM ARQUIVOS ---
blocos_de_dados = [
    ("RESULTADOS - SOMA LINEAR (O(n))", dados_linear),
    ("RESULTADOS - SOMA QUADRÁTICA (O(n))", dados_quadratico),
    ("RESULTADOS - REDUÇÃO LOGARÍTMICA (O(log n))", dados_logaritmico)
]

salvar_resultados_excel(*blocos_de_dados)
salvar_resultados_pdf(info_pc, *blocos_de_dados)