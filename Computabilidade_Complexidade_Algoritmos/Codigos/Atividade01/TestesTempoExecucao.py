import time
import platform
import psutil
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

try:
    import cpuinfo
except ImportError:
    cpuinfo = None

# Função foi renomeada para clareza
def somar_linear(n):
    """Calcula a soma de 1 até n."""
    soma = 0
    for i in range(1, n + 1):
        soma += i
    return soma

# Nova função para a soma quadrática
def somar_quadratico(n):
    """Calcula a soma dos quadrados de 1 até n."""
    soma = 0
    for i in range(1, n + 1):
        soma += i**2  # A única mudança é o cálculo do quadrado (i**2)
    return soma

def get_processador():
    if cpuinfo:
        info = cpuinfo.get_cpu_info()
        return info.get('brand_raw', 'Não identificado')
    
    proc = platform.processor()
    if not proc:
        proc = platform.uname().processor
    if not proc:
        proc = "Não identificado"
    return proc

def info_sistema():
    print("\n--- ESPECIFICAÇÕES DO SISTEMA ---")
    print(f"Sistema Operacional: {platform.system()} {platform.release()} ({platform.version()})")
    print(f"Processador: {get_processador()}")
    print(f"Núcleos Físicos: {psutil.cpu_count(logical=False)}")
    print(f"Núcleos Lógicos: {psutil.cpu_count(logical=True)}")
    print(f"Memória Total: {round(psutil.virtual_memory().total / (1024**3), 2)} GB")
    print("---------------------------------\n")

def salvar_resultados_excel(titulo_bloco_1, dados_bloco_1, titulo_bloco_2, dados_bloco_2):
    """Salva os resultados de ambos os testes em uma planilha Excel."""
    try:
        # Tenta carregar um arquivo existente ou cria um novo
        try:
            workbook = openpyxl.load_workbook("resultados_benchmark.xlsx")
            sheet = workbook.active
            # Adiciona uma linha em branco para separar dos testes anteriores, se houver
            if sheet.max_row > 1:
                sheet.append([])
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Resultados"

        # --- Bloco 1: Soma Linear ---
        # Adiciona o título do primeiro bloco de testes
        cell_titulo_1 = sheet.cell(row=sheet.max_row + 1, column=1, value=titulo_bloco_1)
        cell_titulo_1.font = Font(bold=True)
        sheet.merge_cells(start_row=cell_titulo_1.row, start_column=1, end_row=cell_titulo_1.row, end_column=3)
        
        # Adiciona os cabeçalhos
        headers = ["TESTES", "NUMERO DE PASSOS", "TEMPO DE EXECUÇÃO"]
        sheet.append(headers)
        
        # Adiciona os dados
        for resultado in dados_bloco_1:
            sheet.append(resultado)

        # Adiciona uma linha em branco para separar os blocos
        sheet.append([]) 

        # --- Bloco 2: Soma Quadrática ---
        # Adiciona o título do segundo bloco de testes
        cell_titulo_2 = sheet.cell(row=sheet.max_row + 1, column=1, value=titulo_bloco_2)
        cell_titulo_2.font = Font(bold=True)
        sheet.merge_cells(start_row=cell_titulo_2.row, start_column=1, end_row=cell_titulo_2.row, end_column=3)
        
        # Adiciona os cabeçalhos novamente
        sheet.append(headers)

        # Adiciona os dados
        for resultado in dados_bloco_2:
            sheet.append(resultado)
            
        # Ajusta a largura das colunas para melhor visualização
        for col in range(1, 4):
            sheet.column_dimensions[get_column_letter(col)].width = 25

        # Salva o arquivo
        workbook.save("resultados_benchmark.xlsx")
        print("\nResultados de ambos os testes foram salvos com sucesso em 'resultados_benchmark.xlsx'")
    except Exception as e:
        print(f"Ocorreu um erro ao salvar a planilha: {e}")

# --- Programa Principal ---
qtde_testes = int(input("Quantos testes deseja realizar para cada tipo de soma? "))
info_sistema()

# --- 1. TESTES DE SOMA LINEAR ---
print("--- INICIANDO TESTES: SOMA LINEAR ---")
dados_linear = []
tempos_linear = []

for teste in range(1, qtde_testes + 1):
    n = teste * 100
    inicio = time.perf_counter()
    resultado = somar_linear(n)
    fim = time.perf_counter()
    tempo = fim - inicio
    tempos_linear.append(tempo)
    print(f"Teste {teste}: Somando até {n} → Resultado: {resultado} | Tempo: {tempo:.6f} segundos")
    dados_linear.append([teste, n, tempo])

print("\n--- RESUMO (Soma Linear) ---")
print(f"Tempo total: {sum(tempos_linear):.6f} segundos")
print(f"Tempo médio por teste: {sum(tempos_linear) / qtde_testes:.6f} segundos")
print("----------------------------\n")

# --- 2. TESTES DE SOMA QUADRÁTICA ---
print("--- INICIANDO TESTES: SOMA QUADRÁTICA ---")
dados_quadratico = []
tempos_quadratico = []

for teste in range(1, qtde_testes + 1):
    n = teste * 100
    inicio = time.perf_counter()
    resultado = somar_quadratico(n)
    fim = time.perf_counter()
    tempo = fim - inicio
    tempos_quadratico.append(tempo)
    print(f"Teste {teste}: Soma quadrática até {n} → Resultado: {resultado} | Tempo: {tempo:.6f} segundos")
    dados_quadratico.append([teste, n, tempo])

print("\n--- RESUMO (Soma Quadrática) ---")
print(f"Tempo total: {sum(tempos_quadratico):.6f} segundos")
print(f"Tempo médio por teste: {sum(tempos_quadratico) / qtde_testes:.6f} segundos")
print("--------------------------------\n")


# --- 3. SALVAR TUDO NO EXCEL ---
salvar_resultados_excel(
    "RESULTADOS CÓDIGO 1 - SOMA LINEAR",
    dados_linear,
    "RESULTADOS CÓDIGO 2 - SOMA QUADRATICA",
    dados_quadratico
) 